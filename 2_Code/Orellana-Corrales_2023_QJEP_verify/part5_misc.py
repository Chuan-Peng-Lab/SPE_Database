# -*- coding: utf-8 -*-
"""Part 5: codebook check, per-group gender, paper email, JSON vs CSV field comparison."""
import pandas as pd
import openpyxl
import json, os

BASE = '/Users/hcp4715/Downloads/Collaborations/SPE_Database/SPE_Database/1_Data/Orellana-Corrales_2023_QJEP'

# ---- Codebook ----
wb = openpyxl.load_workbook(os.path.join(BASE, 'Codebook_Orellana-Corrales_2023_QJEP_Exp1_Clean.xlsx'), read_only=True)
print('codebook sheets:', wb.sheetnames)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb[wb.sheetnames[0]]
cb = list(ws.iter_rows(values_only=True))
print('codebook rows (incl header):', len(cb))
for r in cb:
    print('  ', r)

clean = pd.read_csv(os.path.join(BASE, 'Orellana-Corrales_2023_QJEP_Exp1_Clean.csv'))
print()
print('Clean columns (%d):' % len(clean.columns), list(clean.columns))
print('codebook Variable_name set == Clean columns set:',
      set(r[0] for r in cb[1:]) == set(clean.columns))

# ---- per-group gender from subj_info ----
subj = pd.read_csv(os.path.join(BASE, 'Orellana-Corrales_2023_QJEP_Exp1_subj_info.csv'))
print()
print('subj_info per-group Gender counts:')
print(subj.groupby(['Group','Gender']).size().to_string())
print('subj_info per-group total:')
print(subj.Group.value_counts().to_dict())

# ---- JSON fields ----
pj = json.load(open(os.path.join(BASE, 'Orellana-Corrales_2023_QJEP.json')))
ej = json.load(open(os.path.join(BASE, 'Orellana-Corrales_2023_QJEP_Exp1.json')))
print()
print('paper JSON keys:', list(pj.keys()))
print('exp JSON keys:', list(ej['exp1'].keys()))
print('paper JSON Email:', pj.get('Email'), '| Country:', pj.get('Country'), '| City:', pj.get('City'))
print('exp JSON Block_Structure:', ej['exp1']['Block_Structure'])
print('exp JSON Trial_Structure:', ej['exp1']['Trial_Structure'])
print('exp JSON Stimulus_Properties:', ej['exp1']['Stimulus_Properties'])
print('exp JSON Physical_Environment:', ej['exp1']['Physical_Environment'])
print('exp JSON Experimental_Design:', ej['exp1']['Experimental_Design'])
print('exp JSON Collected_date:', ej['exp1']['Collected_date'])
